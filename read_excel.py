import openpyxl
import json

path = r"C:\Users\Kuaige\Desktop\ERP_OA_CRM_WMS_IoT_PLM整合梳理表.xlsx"

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append([str(cell) if cell is not None else "" for cell in row])
        data[sheet_name] = sheet_data
    print(json.dumps(data, ensure_ascii=False, indent=2))
except Exception as e:
    print(f"Error: {e}")
