"""
Excel报表生成引擎模块

使用openpyxl生成Excel报表文件。

Author: Luigi Lu
Date: 2025-01-15
"""

from typing import Dict, Any, List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger


class ExcelEngine:
    """
    Excel报表生成引擎

    使用openpyxl生成Excel文件。
    """

    def generate(self, config: dict[str, Any], data: dict[str, Any]) -> BytesIO:
        """
        生成Excel报表

        Args:
            config: 报表配置
            data: 报表数据

        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = config.get("name", "报表")

        # 设置中文字体
        chinese_font = Font(name="Microsoft YaHei", size=11)

        # 渲染组件
        components = config.get("components", [])
        for component in components:
            self._render_component(ws, component, data, chinese_font)

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _render_component(
        self,
        ws,
        component: dict[str, Any],
        data: dict[str, Any],
        font: Font,
    ):
        """
        渲染单个组件

        Args:
            ws: 工作表对象
            component: 组件配置
            data: 数据
            font: 字体
        """
        component_type = component.get("type")
        x = component.get("x", 0)
        y = component.get("y", 0)

        if component_type == "table":
            self._render_table(ws, component, data, font, x, y)
        elif component_type == "text":
            self._render_text(ws, component, font, x, y)
        elif component_type == "chart":
            # Excel图表需要特殊处理，这里简化实现
            logger.warning("Excel图表渲染暂未实现")
        elif component_type == "image":
            # Excel图片需要特殊处理，这里简化实现
            logger.warning("Excel图片渲染暂未实现")

    def _render_table(
        self,
        ws,
        component: dict[str, Any],
        data: dict[str, Any],
        font: Font,
        start_x: int,
        start_y: int,
    ):
        """
        渲染表格组件

        Args:
            ws: 工作表对象
            component: 组件配置
            data: 数据
            font: 字体
            start_x: 起始X坐标（列）
            start_y: 起始Y坐标（行）
        """
        # 获取表格数据
        table_data = data.get(component.get("data_source", ""), [])
        columns = component.get("columns", [])

        # 渲染表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        
        for col_idx, col in enumerate(columns):
            cell = ws.cell(row=start_y + 1, column=start_x + col_idx + 1)
            cell.value = col.get("title", col.get("dataIndex", ""))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 渲染数据行
        for row_idx, row_data in enumerate(table_data):
            for col_idx, col in enumerate(columns):
                cell = ws.cell(row=start_y + row_idx + 2, column=start_x + col_idx + 1)
                data_index = col.get("dataIndex", "")
                cell.value = row_data.get(data_index, "")
                cell.font = font
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 自动调整列宽
        for col_idx, col in enumerate(columns):
            column_letter = get_column_letter(start_x + col_idx + 1)
            ws.column_dimensions[column_letter].width = 15

    def _render_text(
        self,
        ws,
        component: dict[str, Any],
        font: Font,
        start_x: int,
        start_y: int,
    ):
        """
        渲染文本组件

        Args:
            ws: 工作表对象
            component: 组件配置
            font: 字体
            start_x: 起始X坐标（列）
            start_y: 起始Y坐标（行）
        """
        content = component.get("content", "")
        text_type = component.get("textType", "paragraph")

        cell = ws.cell(row=start_y + 1, column=start_x + 1)
        cell.value = content

        if text_type == "title":
            level = component.get("level", 1)
            cell.font = Font(name="Microsoft YaHei", size=14 + (6 - level), bold=True)
        else:
            cell.font = font

        cell.alignment = Alignment(horizontal="left", vertical="center")

